# -*- coding: utf-8 -*-
"""
Created on Fri May 26 06:54:39 2023

@author: Mario
"""

import openpyxl

class Funexcel():
    def __init__(self,driver):
        self.driver =driver
        
    def getRowCount(file,path,sheetName):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file,sheetName):
        Worbook=openpyxl.load_workbook(file)
        sheet=Worbook[sheetName]
        return (sheet.max_column)
    
    def readData(file,path,sheetName,rownum,columno):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName]
        return sheet.cell(row=rownum, column=columno).value
    
    def writeData(file,path,sheetName,rownum,columno,data):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName]
        sheet.cell(row=rownum, column=columno).value=data
        Worbook.save(path)